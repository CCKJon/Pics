from urllib.parse import urlparse
import openpyxl
import webbrowser
import requests
import os
import time

# Open the xlsx file
workbook = openpyxl.load_workbook('bou+pro+0922_entries.xlsx')

# Select the worksheet by name
worksheet = workbook['Sheet1']

i = 0

# Loop through each row in the worksheet
for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, values_only=True):
    if row[12] != "":
        i = i+1
        # Access the data in each column of the row
        idcolumn = row[0]+1
        column1 = row[8]
        column2 = row[9]
        column3 = row[12]
        # [0:(len(row[12]))-4]
        # Do something with the data
        line = "".join((column1 + "-" + column2 + "_" + "{:03d}".format(i)).split())
        new_line = "BOU-PRO-0922_BoulderJunction_092022_" + line
        
        # Get the hyperlink from cell
        # print(idcolumn)
        # int(idcolumn)
        hyperlink = worksheet.cell(row=12, column=13).hyperlink.target
        webbrowser.open(hyperlink)
        # filename = hyperlink.display
        linkaddress = hyperlink
        old_name = column3
        new_name = new_line + '.jpg'
        time.sleep(9)
        os.rename(old_name, new_name)
        break
        # url = hyperlink.target
        # response = requests.get(url)
        # Open the hyperlink in the default web browser


# Close the xlsx file
workbook.close()

